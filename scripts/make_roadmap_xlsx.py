"""Generate ROADMAP_STATUS.xlsx from the markdown roadmap + current state.

Three sheets:
  1. Tasks      — every roadmap task (C-/I-/T-/P-/S-) with status, %, dates
  2. Milestones — week-numbered milestones with Done/Pending/Future
  3. Updates    — recent commit/event log this week

Color coding:
  - Done       → green
  - In Progress→ yellow
  - Pending    → blue
  - Blocked    → red
  - Future     → gray
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).parent.parent
OUT = REPO.parent / "ROADMAP_STATUS.xlsx"


def _resolve_writable_path(p: Path) -> Path:
    """If p is open in Excel (Windows file lock), pick a numbered fallback."""
    if not p.exists():
        return p
    try:
        with open(p, "ab"):
            return p  # writable
    except PermissionError:
        for i in range(1, 50):
            cand = p.with_name(f"{p.stem}_v{i}{p.suffix}")
            if not cand.exists():
                return cand
            try:
                with open(cand, "ab"):
                    return cand
            except PermissionError:
                continue
        raise RuntimeError("no writable filename available")

TODAY = date(2026, 4, 25)
START = date(2026, 4, 23)  # Roadmap kickoff

# Status palette
GREEN  = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="BDD7EE")
RED    = PatternFill("solid", fgColor="FFC7CE")
GRAY   = PatternFill("solid", fgColor="E7E6E6")
HEADER = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(border_style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def fill_for_status(s: str) -> PatternFill:
    s = s.lower()
    if s.startswith("done"):     return GREEN
    if "progress" in s or "running" in s or "active" in s: return YELLOW
    if s.startswith("pending") or s.startswith("ready"):   return BLUE
    if s.startswith("blocked") or s.startswith("at risk"): return RED
    return GRAY


# ── Task data — one row per roadmap item ─────────────────────────────────────

TASKS = [
    # (Code, Task, Phase, Status, %Complete, OrigTimeline, ActualStart, ActualEnd, Depends, Notes)
    # ── Phase 0: Initial Physics Stack (project foundation, weeks before audit) ─
    ("H-0", "Initial commit: PlasmaNet neural surrogate scaffolding", "Foundation", "Done",
     100, "Project start", "Pre-2026-04", "Pre-2026-04", "—",
     "Repo bootstrap, package layout, training harness skeleton"),
    ("H-1", "Initial physics: Cantera equilibrium, Saha, std atmosphere, plasma freq", "Foundation", "Done",
     100, "Week 1", "Pre-2026-04", "Pre-2026-04", "H-0",
     "First analytical pipeline; equilibrium-only ne prediction"),
    ("H-2", "PlasmaNet v1 NN (4-input stagnation-only surrogate)", "Foundation", "Done",
     100, "Week 1-2", "Pre-2026-04", "Pre-2026-04", "H-1",
     "Inputs: Mach, altitude, nose_radius, half_angle. Output: stagnation ne"),
    ("H-3", "Continuous training loop + architecture-aware checkpoints", "Foundation", "Done",
     100, "Week 2", "Pre-2026-04", "Pre-2026-04", "H-2",
     "Auto-resumes training; 45K data points by round 2"),
    ("H-4", "SBIR demo: self-contained launcher with interactive UI", "Foundation", "Done",
     100, "Week 2", "Pre-2026-04", "Pre-2026-04", "H-2",
     "First customer-demo polish, predates SimOps integration"),

    # ── Phase 1: CFD batch + DRGEP chemistry analysis ──────────────────────
    ("H-5", "Parametric geometry generator (sphere-cone family)", "CFD Pipeline", "Done",
     100, "Week 3", "Pre-2026-04", "Pre-2026-04", "—",
     "CadQuery STEP output for 5 nose-radius/half-angle combos"),
    ("H-6", "5 geometries meshed + 40 SU2 Euler configs (5 geom x 8 flight)", "CFD Pipeline", "Done",
     100, "Week 3", "Pre-2026-04", "Pre-2026-04", "H-5",
     "Foundation for CFD-derived training data"),
    ("H-7", "CFD batch runner + GCP execution", "CFD Pipeline", "Done",
     100, "Week 3-4", "Pre-2026-04", "Pre-2026-04", "H-6",
     "Ran 40 Euler cases on openfoam-hgv VM"),
    ("H-8", "DRGEP transient 0D reactor analysis (R2 dominant chemistry)", "Chemistry", "Done",
     100, "Week 3-4", "Pre-2026-04", "Pre-2026-04", "H-1",
     "Reduced 11-species mechanism; condition-adaptive map across 2000-20000K"),
    ("H-9", "PlasmaNet v2 model + validated CFD extraction pipeline", "CFD Pipeline", "Done",
     100, "Week 4", "Pre-2026-04", "Pre-2026-04", "H-7",
     "Retrained on Euler-derived data; bigger architecture"),

    # ── Phase 2: Audit + Physics Stack Rebuild (the 'big rewrite') ─────────
    ("H-10", "Audit: surfaced 6 critical/minor physics issues", "Audit", "Done",
     100, "2026-04-21/22", "2026-04-21", "2026-04-22", "H-9",
     "Pitot vs isentropic, NEQ contamination, activation E units, DRGEP duplicates, partition fns, atmosphere extent"),
    ("H-11", "Audit fixes: pitot, NEQ flag, units, single-source DRGEP, NIST partitions", "Audit", "Done",
     100, "2026-04-22", "2026-04-22", "2026-04-22", "H-10",
     "All 6 audit findings resolved; commit 209e0ed"),
    ("H-12", "Post-audit physics stack: pitot + wave + LOS + UQ + RAM-C harness", "Physics Stack", "Done",
     100, "2026-04-22/23", "2026-04-22", "2026-04-23", "H-11",
     "0d682cc: full electromagnetic + chemistry-UQ pipeline; canonical RAM-C benchmark harness"),
    ("H-13", "CFD field extractor + end-to-end detectability example", "Physics Stack", "Done",
     100, "2026-04-23", "2026-04-23", "2026-04-23", "H-12",
     "08cfa7a: read SU2 Euler VTU -> ne field -> aspect-resolved attenuation"),
    ("H-14", "T-corrected Euler (real-gas) CFD post-processing", "Physics Stack", "Done",
     100, "2026-04-23", "2026-04-23", "2026-04-23", "H-13",
     "6ae158b: bridges equilibrium ne until coupled-chem CFD lands"),
    ("H-15", "Post-audit project overview + Path C/SimOps roadmap docs", "Planning", "Done",
     100, "2026-04-23", "2026-04-23", "2026-04-23", "H-12",
     "484e07f: PROJECT_OVERVIEW_POST_AUDIT.md + ROADMAP_SIMOPS_INTEGRATION.md"),

    # ── Path C: SU2-NEMO unblock + RAM-C validation ─────────────────────────
    ("C-0", "Inventory SU2-NEMO + Mutation++ on GCP VM", "Path C", "Done",
     100, "Day 0.5", "2026-04-23", "2026-04-23", "H-15",
     "Mutation++ at /opt/su2-nemo/, MPP_DATA_DIRECTORY found"),
    ("C-1", "Resolve SU2-NEMO segfault (FLUID_MODEL= SU2_NONEQ)", "Path C", "Done",
     100, "Day 1-3", "2026-04-23", "2026-04-23", "C-0",
     "Root cause: missing FLUID_MODEL. See SU2_NEMO_FIX.md"),
    ("C-2", "Validate NEMO on blunt_cone M10 @ 30 km", "Path C", "Done",
     100, "Day 4-5", "2026-04-23", "2026-04-23", "C-1",
     "T_tr=5669K, T_ve=3948K, ne=3.0e18 m^-3 (real 2-T NEQ)"),
    ("C-3", "Validate NEMO on RAM-C II 61 km M22.5 (first pass, 63k mesh)", "Path C", "Done",
     100, "Day 6-7", "2026-04-23", "2026-04-24", "C-2",
     "log10 err +1.08 (top-50 robust), all BLACKOUT bands match J&C"),
    ("C-3a", "Refined-mesh ramp (2.67M tets, M10->M15->M18->M22.5)", "Path C", "In Progress",
     65, "Day 7-8", "2026-04-24", "2026-04-25 ETA", "C-3",
     "M10/M15 done; M18 iter 147/200 Rho_0=-1.9; M22.5 next"),
    ("C-3b", "Run validate_ram_c_nemo on refined M22.5 result", "Path C", "Pending",
     0, "Day 8", "2026-04-25 ETA", "—", "C-3a",
     "Will fire on M22.5 completion via monitor"),
    ("C-3c", "AIR-11 (Mutation++) ramp variant — fallback if AIR-5 misses", "Path C", "Pending",
     90, "Day 8-9 (if needed)", "—", "—", "C-3b",
     "Pre-staged: scripts/ram_c_refined_ramp_air11.sh ready"),
    ("C-4", "Port 40-case CFD batch to NEMO", "Path C", "Pending",
     30, "Week 2 end", "—", "—", "C-3b",
     "Scaffolding ready: plasmanet/nemo_config.py + generate_nemo_batch.py"),
    ("S-1", "Streamline-based chemistry fallback (Path A)", "Path A", "Done",
     100, "Deprecated", "—", "—", "—",
     "Path C succeeded — Path A fallback not needed"),

    # ── Infrastructure: SimOps integration on AWS ───────────────────────────
    ("I-1", "PlasmaNet Service Fargate stack (CDK)", "SimOps Infra", "Done",
     100, "Week 2-3", "2026-04-24", "2026-04-25", "C-1",
     "cdk/plasmanet_service_stack.py — ECR + ALB + Fargate + SSM, cdk synth clean"),
    ("I-1a", "Mock server (FastAPI) for local + contract validation", "SimOps Infra", "Done",
     100, "—", "2026-04-24", "2026-04-25", "—",
     "plasmanet/mock_server.py + 27 contract tests; Windows UTF-8 fixed"),
    ("I-2", "SU2-NEMO Worker Batch stack (CDK)", "SimOps Infra", "Done",
     100, "Week 2-3", "2026-04-24", "2026-04-25", "C-1",
     "cdk/plasmanet_worker_stack.py — Spot+OD CE, S3 lifecycle, EventBridge->Lambda"),
    ("I-2a", "Worker Dockerfile (placeholder for SU2-NEMO base)", "SimOps Infra", "In Progress",
     30, "Week 3", "2026-04-25", "—", "I-2",
     "Multi-stage Dockerfile shipped; real SU2-NEMO base image pending"),
    ("I-2b", "Lambda webhook (UUID resolver + DescribeJobs fallback)", "SimOps Infra", "Done",
     100, "Week 3", "2026-04-25", "2026-04-25", "I-2",
     "lambda/simulation_complete + 3 moto integration tests + IAM grant"),
    ("I-3", "/api/plasma/analyze endpoint", "SimOps Infra", "Done",
     100, "Week 3-4", "2026-04-24", "2026-04-25", "I-1a",
     "Mock returns DetectabilityReport with stagnation+UQ+aspect_scan"),
    ("I-4", "/api/plasma/benchmark/ram_c endpoint", "SimOps Infra", "Done",
     100, "Week 3-4", "2026-04-24", "2026-04-25", "I-1a",
     "Returns 12-case J&C comparison table (4 alts x 3 freqs)"),
    ("I-5", "Frontend aspect polar plot + UQ band", "SimOps Infra", "Done",
     100, "Week 4-5", "2026-04-24", "2026-04-25", "I-3",
     "frontend/ React+Vite+Tailwind+TS, LIVE/MOCK badge fallback"),
    ("I-5a", "Frontend altitude/Mach selectors + station profile chart", "SimOps Infra", "Done",
     100, "Week 5", "2026-04-25", "2026-04-25", "I-5",
     "Shipped commit eadcbc8: button-row selectors, 2nd chart with synthetic decay until CFD lands"),
    ("I-5b", "Lambda test coverage gaps (tags / exception / missing jobId)", "SimOps Infra", "Done",
     100, "—", "2026-04-25", "2026-04-25", "I-2b",
     "Shipped commit 4a6f8a0: 6 lambda tests total now, all green"),
    ("I-5c", "Frontend Storybook + RTL component tests + CI frontend job", "SimOps Infra", "Done",
     100, "Week 5", "2026-04-25", "2026-04-25", "I-5",
     "5 commits 365de52..118845e: Storybook 10, extracted LiveMockBadge+FlightSelectors, vitest RTL, ci frontend job"),
    ("I-9", "/api/plasma/report PDF endpoint + auto RAM-C benchmark attribution", "SimOps Infra", "Done",
     100, "Week 4", "2026-04-25", "2026-04-25", "I-3",
     "Commits 27cf8f0 + 5faff1d + b422c21: A4 one-pager with explicit canonical-point attribution"),
    ("I-10", "Agent tools: analyze_plasma + generate_plasma_report (Pydantic AI)", "SimOps Infra", "Done",
     100, "Week 4", "2026-04-25", "2026-04-25", "I-3",
     "Commits 6368146 + a6db8b1: 13 agent tests; soft pydantic_ai import for KhoriumAgents"),
    ("I-11", "Repo cleanup: dead modules, legacy/, gitignore, pyproject extras", "SimOps Infra", "Done",
     100, "—", "2026-04-25", "2026-04-25", "—",
     "bd69e20: deleted 1068 lines of dead code, archived pre-React demo, fixed pyproject"),

    # ── Closing the log10 error gap (post-M22.5 ramp 2 result) ────────────
    ("F-1", "AIR-11 chemistry ramp — primary fix for downstream plasma collapse", "Validation Fix", "Done",
     100, "Day 8 (overnight)", "2026-04-25", "2026-04-25", "C-3a",
     "EXHAUSTED. 6 distinct attempts (cold-start, seeded ions, charge-balanced restart, species floor, smaller seed) all NaN-frozen. Root cause: Mutation++ EOS conventions differ from CSU2TCLib (built-in NEMO) — restart values produce non-physical (T<50K or NaN) state in 89%+ cells. Cannot fix at cfg/restart layer in v7.5.1. Pivoted to AIR-7 (F-8). Documented in commits 6b72f33, 65b105d, 5cababc, ac73c94, fcb9bdd. The negative result is itself a research finding."),
    ("F-2", "AIR-11 + non-catalytic Navier-Stokes combined ramp (fallback)", "Validation Fix", "Done",
     100, "Day 9", "—", "2026-04-25", "F-1",
     "DELETED — superseded by F-8 (AIR-7 pivot). Catalytic-wall fix doesn't apply since AIR-11 itself is unworkable in v7.5.1."),
    ("F-3", "Anisotropic near-wall mesh refinement (BL layers)", "Validation Fix", "Pending",
     20, "Day 10-11", "2026-04-25", "—", "F-9",
     "Script written (mesh_ram_c_BL.py) but blocked on F-9 MPI rebuild — current serial SU2-NEMO can't handle 1.5-2M anisotropic cells in reasonable time. Re-prioritize after MPI binary available."),
    ("F-4", "Tight-convergence rerun of existing M22.5 (cheap polish)", "Validation Fix", "Pending",
     0, "Day 8 (cheap, opportunistic)", "—", "—", "F-8",
     "Re-run AIR-7 M22.5 at tighter convergence after F-8 finishes. Skip if F-8 lands |log10 err| < 0.3 already."),
    ("F-5", "T_wall sensitivity sweep (1500/2000/2500/3000 K)", "Validation Fix", "Pending",
     0, "Day 11-12", "—", "—", "F-8 + F-9",
     "T_wall=2500K (RAM-C SiO2 ablation temp) is a calibration knob. Real value varies 1500-3000K. Sweep on F-8 result. ~4×4h wall time WITHOUT MPI; ~4×30min with F-9 MPI rebuild — much more attractive after MPI."),
    ("F-6", "ParaView 3D diagnostic — visualize plasma collapse geometrically", "Validation Fix", "Done",
     100, "Day 8 (any time)", "2026-04-25", "2026-04-25", "C-3a",
     "Built scripts/paraview_3d_diagnostic.py (pyvista-based, off-screen). Generates 5 figures from any AIR-5/AIR-7 VTU: ne isosurface, ne+Mach+T_tr slices on y=0, axial sheath profile with J&C stations marked. AIR-5 baseline figures saved to data/checkpoints/air5_M22_5_A61/. Axial profile is the smoking gun — shows AIR-5 ne dropping 4 orders by z/L=0.14 vs J&C published persistent plasma."),
    ("F-7", "Direct LOS attenuation comparison vs J&C published dB", "Validation Fix", "Done",
     100, "Day 8 (parallel with F-1)", "2026-04-25", "2026-04-25", "—",
     "DONE in commit 64906f9. Added db_margin_to_published() to validate_ram_c_nemo.py. AIR-5 result: VHF margins +29/+52 dB CONSISTENT with J&C BLACKOUT, X-band -15 dB INCONSISTENT (we predict DEGRADED, J&C says BLACKOUT). Comparison script (compare_air5_vs_air7_ramc.py) extends this to side-by-side AIR-5 vs AIR-7."),

    # ── New tasks from AIR-7 pivot (post-AIR-11 exhaustion) ────────────────
    ("F-8", "AIR-7 ramp (built-in CSU2TCLib, e- + NO+) — primary path", "Validation Fix", "In Progress",
     50, "Day 8-9", "2026-04-25", "2026-04-26 ETA", "C-3a",
     "Pivoted from F-1 after AIR-11 exhausted. AIR-7 species [e-, N2, O2, NO, N, O, NO+] uses same EOS as AIR-5 (no thermo mismatch), supports EULER_IMPLICIT (10x faster than AIR-11 EXPLICIT), provides direct CFD ne. Currently running v7 (CFL=0.2 fixed): M10 stage hit chemistry stiffness wall at iter 234, killed and resumed via ram_c_air7_resume_from_m10.sh. M15 in progress at iter ~160/250 (RhoU limit-cycle at -0.79 expected). M18+M22.5 to follow. Expected log10 err -0.5 to -1.0 (vs AIR-5 -1.59). Total ETA ~13h serial."),
    ("F-9", "SU2-NEMO MPI rebuild (parallel speedup 10-15x)", "Validation Fix", "In Progress",
     20, "Day 8-9", "2026-04-25", "2026-04-26 ETA", "—",
     "Delegated to second Claude instance. Build SU2-NEMO v7.5.1 with -Dwith-mpi=enabled in separate prefix /opt/su2-nemo-mpi/, install, verify with mpirun -n 4 smoke test. Critical safety: don't touch running SU2_CFD or modify /opt/su2-nemo/. Expected impact: 50s/iter -> 4s/iter, 1500-iter stage 21h -> 1.7h. Unblocks F-3 (BL mesh), F-5 (T_wall sweep), and any future Khorium hypersonics work."),
    ("F-10", "AIR-5 vs AIR-7 vs J&C 1972 comparison report", "Validation Fix", "In Progress",
     50, "Day 9 (when F-8 done)", "2026-04-25", "—", "F-8",
     "Pipeline pre-baked: scripts/compare_air5_vs_air7_ramc.py runs validate_ram_c_nemo on both VTUs, builds side-by-side markdown table (per-station ne, headline log10 err, dB margins per band). Drops final report the moment F-8 M22.5 finishes."),
    ("F-11", "AIR-7 chemistry development trajectory plot", "Validation Fix", "Done",
     100, "Day 8", "2026-04-25", "2026-04-25", "F-8",
     "scripts/plot_air7_chemistry_dev.py: pulls v7 history.csv from VM, plots species + bulk residuals with limit-cycle annotation. Confirmed chemistry on at iter 2, RhoU plateau at -0.77. Diagnostic for paper figure."),

    ("I-6", "Agent tool: analyze_plasma(condition)", "SimOps Infra", "Pending",
     0, "Week 4", "—", "—", "I-3",
     "KhoriumAgents Pydantic AI tool binding"),
    ("I-7", "GitHub Actions CI workflow", "SimOps Infra", "Done",
     100, "—", "2026-04-25", "2026-04-25", "—",
     ".github/workflows/ci.yml — cdk synth + 33 contract tests"),
    ("I-8", "SimOps Integration design doc", "SimOps Infra", "Done",
     100, "—", "2026-04-23", "2026-04-23", "—",
     "docs/SIMOPS_INTEGRATION.md — 398 lines, two-layer architecture"),

    # ── Training: NN surrogate updates ──────────────────────────────────────
    ("T-1", "Retrain PlasmaNet NN on NEMO-derived ne fields", "Training", "Pending",
     0, "Week 4-5", "—", "—", "C-4",
     "Blocked on C-4 (40-case batch) for training data"),
    ("T-2", "Field-NN: learn ne(x,y,z) from NEMO output", "Training", "Pending",
     0, "Week 5-7", "—", "—", "C-4",
     "Geometry-aware extension; future research track"),

    # ── Production polish ───────────────────────────────────────────────────
    ("P-1", "AFRL SBIR demo — live web walkthrough", "Production", "Pending",
     0, "Week 6-8", "—", "—", "I-5",
     "Demo readiness: needs frontend selectors + real M22.5 numbers"),
    ("P-2", "Auto-generated PDF plasma report", "Production", "Pending",
     0, "Week 7-8", "—", "—", "I-5",
     "Polar plot + envelope + UQ + RAM-C self-check"),
    ("P-3", "Production billing — metered per analyze call", "Production", "Pending",
     0, "Week 8-9", "—", "—", "I-3",
     ""),
    ("P-4", "Paper revision — AIAA Journal of Thermophysics", "Production", "In Progress",
     20, "Week 8-10", "2026-04-24", "—", "C-4",
     "Other instance updated draft; awaiting M22.5 result for figures"),
    ("P-5", "Geometry-aware field NN (publication target)", "Research", "Future",
     0, "Week 10+", "—", "—", "T-2",
     "Long-term research direction"),
]


# ── Milestone data — week-anchored ─────────────────────────────────────────

MILESTONES = [
    # (Week, Milestone, Original Target, Actual/Projected, Status, Notes)
    ("Week 1 end", "NEMO segfault fixed, blunt_cone M10 passes",
     "2026-04-30", "2026-04-23", "Done",
     "Hit on Day 1, 7 days early. SU2_NONEQ fix unblocked everything."),
    ("Week 2 end", "RAM-C 61 km NEMO prediction within factor of 2 of measured",
     "2026-05-07", "2026-04-25 ETA", "In Progress",
     "First-pass log10 err +1.08; refined-mesh M22.5 result tonight"),
    ("Week 3 end", "40-case batch ported to NEMO with aspect-resolved reports",
     "2026-05-14", "2026-05-08 projected", "Pending",
     "Scaffolding ready (nemo_config.py); blocked on C-3 success"),
    ("Week 4 end", "SimOps /api/plasma/analyze live on staging",
     "2026-05-21", "2026-05-01 projected", "In Progress",
     "Mock server + CDK both done; missing real Fargate deploy"),
    ("Week 5 end", "Frontend polar plot + envelope UI shipped",
     "2026-05-28", "2026-04-26 ETA", "In Progress",
     "Polar plot done; selectors + 2nd chart in queue"),
    ("Week 8 end", "AFRL SBIR demo ready",
     "2026-06-18", "2026-06-18", "Pending",
     "On-track; needs P-1 walkthrough preparation"),
    ("Week 10 end", "Paper submitted",
     "2026-07-02", "2026-07-02", "Pending",
     "Needs full 40-case NEMO batch + paper revision (P-4)"),
]


# ── Recent updates log (this week) ──────────────────────────────────────────

UPDATES = [
    # (Date, Commit, Description)
    # ── Pre-audit foundation (chronological from git log --reverse) ─────────
    ("Pre-04-23", "a0e45ef", "Initial commit: PlasmaNet neural surrogate for hypersonic plasma prediction"),
    ("Pre-04-23", "e2af5d7", "Continuous training loop + architecture-aware checkpoints"),
    ("Pre-04-23", "502bccd", "Training round 2: 20 more iterations, 45K total data points"),
    ("Pre-04-23", "04711c7", "SBIR demo: self-contained launcher with interactive UI"),
    ("Pre-04-23", "0183082", "Parametric geometry generator + working DRGEP implementation"),
    ("Pre-04-23", "82eb119", "DRGEP condition-adaptive mechanism map"),
    ("Pre-04-23", "546a7ba", "Corrected DRGEP transient 0D reactor results"),
    ("Pre-04-23", "5d8a118", "5 geometries meshed + 40 SU2 configs ready for GCP"),
    ("Pre-04-23", "51da750", "CFD batch runner + 5 geometries + 40 SU2 configs"),
    ("Pre-04-23", "b4603b7", "Complete DRGEP map (2000-20000K) + CFD extractor + Dockerfile"),
    ("Pre-04-23", "bfc8e02", "PlasmaNet v2 model + validated CFD extraction pipeline"),
    ("2026-04-22", "0f70e3e", "Rewrite README as authoritative source of truth for auditing"),
    ("2026-04-22", "d7ea0cc", "Clean PlasmaNet v2: retrained from scratch on pure equilibrium data"),
    ("2026-04-22", "209e0ed", "Fix all audit findings: 4 critical + 2 minor"),
    ("2026-04-23", "0d682cc", "Proper physics stack: pitot + wave propagation + LOS + UQ + RAM-C"),
    ("2026-04-23", "08cfa7a", "CFD field extraction + end-to-end detectability example"),
    ("2026-04-23", "6ae158b", "cfd_field: real-gas T correction for SU2 Euler output"),
    ("2026-04-23", "484e07f", "docs: post-audit project overview + Path C/SimOps roadmap"),
    # ── Path C kickoff and execution ────────────────────────────────────────
    ("2026-04-23", "Day 1", "Path C kickoff — SU2-NEMO segfault root-caused (FLUID_MODEL= SU2_NONEQ); first 2-T NEQ result on blunt cone"),
    ("2026-04-23", "ef1f793", "RAM-C II geometry/mesh generator + validation harness"),
    ("2026-04-23", "3eae9a0", "SimOps integration design doc (398 lines)"),
    ("2026-04-23", "f95c9b3", "Frontend polar attenuation plot scaffold"),
    ("2026-04-24", "72de78a", "RAM-C M22.5 first-pass result (small mesh) — log10 err +1.51 raw / +1.08 robust"),
    ("2026-04-24", "fdc4b0d", "Refined RAM-C mesh + ramp infrastructure"),
    ("2026-04-24", "01025fe", "Validation script: robust peak ne + sheath-shell filter; Notion doc updated"),
    ("2026-04-24", "ae81e14", "AIR-11 ramp variant pre-staged"),
    ("2026-04-24", "b245e08", "Mock server schema drift fixed (radar wrapper + required fields)"),
    ("2026-04-24", "d11b11c", "27 contract tests for mock_server routes"),
    ("2026-04-24", "8805a60", "Windows UTF-8 stdout fix for mock_server"),
    ("2026-04-24", "c254f64", "RamCCaseResult.within_uncertainty -> status_match (with alias)"),
    ("2026-04-24", "3e83889", "PlasmaNetService Fargate CDK stack"),
    ("2026-04-24", "50e93a9", "PlasmaNet Worker Batch CDK stack"),
    ("2026-04-24", "656eef4", "Dockerfile (multi-stage) + cdk/README"),
    ("2026-04-24", "ee1f000", "CDK audit fixes: separate exec role, scoped checkpoint perms, boto3, jobName contract"),
    ("2026-04-24", "79c38eb", "MODEL_S3_KEY -> SSM Parameter Store"),
    ("2026-04-24", "60acd34", "Lambda extraction to lambda/simulation_complete/"),
    ("2026-04-25", "d237082", "Fix ramp prev_dir typo (M10 vs M10_0); add resume script"),
    ("2026-04-25", "21920ba", "Coarsen refined mesh: 4.5M -> 2.7M tets after stuck preprocessing"),
    ("2026-04-25", "e579208", "Phase2 watcher: detect orphan SU2 by cwd not cmdline"),
    ("2026-04-25", "718aa77", "Unified RAM-C ramp script (replaces 6 one-shot scripts)"),
    ("2026-04-25", "7d4a83d", "Paper-figure generator (4 PNGs + drop-in markdown blurb)"),
    ("2026-04-25", "aaf0a7a", "M18 fallback recipes (lowcfl / AUSM+ / M16-M17 intermediate)"),
    ("2026-04-25", "2f4ef4f", "GitHub Actions CI workflow"),
    ("2026-04-25", "efa515e", "Lambda UUID detection + DescribeJobs fallback"),
    ("2026-04-25", "2f3a41b", "Lambda moto integration tests (3 cases)"),
    ("2026-04-25", "49e34f7", "CI: add lambda tests to pytest job"),
    ("2026-04-25", "4a6f8a0", "Lambda coverage gaps: tags fallback + ClientError + missing jobId (6 tests now)"),
    ("2026-04-25", "eadcbc8", "Frontend altitude/Mach selectors + station profile chart + loading/error states"),
    ("2026-04-25", "1881a68", "Bake synthetic station_profile into mock_los.json so offline shows both charts"),
    ("2026-04-25", "27cf8f0", "/api/plasma/report PDF endpoint (A4 one-pager, polar+station, 7 tests)"),
    ("2026-04-25", "6368146", "agent_tools: analyze_plasma Pydantic AI tool (8 mocked tests)"),
    ("2026-04-25", "a4faa5a", "M22.5 result finalizer (one-shot scp+validate+figures+notion+commit)"),
    ("2026-04-25", "25e6017", "Multi-stage ramp evolution figure generator"),
    ("2026-04-25", "5faff1d", "Auto RAM-C benchmark in /report (CANONICAL_RAMC_POINTS dispatcher)"),
    ("2026-04-25", "a6db8b1", "agent_tools: generate_plasma_report PDF-as-bytes tool"),
    ("2026-04-25", "b422c21", "/report Validation: explicit attribution to nearest canonical RAM-C point"),
    ("2026-04-25", "bd69e20", "Cleanup: dead modules deleted, pre-React demo -> legacy/, pyproject extras"),
    ("2026-04-25", "365de52", "frontend: Storybook 10 set up (Vite builder, dark-theme decorator)"),
    ("2026-04-25", "34ff21b", "frontend: extract LiveMockBadge + FlightSelectors from App.tsx"),
    ("2026-04-25", "cc82cb9", "frontend: Storybook stories for all components"),
    ("2026-04-25", "ad60e1a", "frontend: vitest+RTL coverage for App + components"),
    ("2026-04-25", "118845e", "ci: frontend job (build + vitest)"),
    # ── AIR-11 attempt 1-6 (all failed, but documented) ─────────────────────
    ("2026-04-25", "6b72f33", "AIR-11 attempt 1: silent-completion bug fix (CONV_FIELD electrons trap)"),
    ("2026-04-25", "65b105d", "AIR-11 attempt 3: seeded ions + adaptive CFL warmup — still NaN-frozen"),
    ("2026-04-25", "5cababc", "AIR-11 attempt 4: warm-start from converted AIR-5 restart"),
    ("2026-04-25", "ac73c94", "AIR-11 attempts 5+6 (charge balance, species floor) + AIR-7 pivot"),
    # ── AIR-7 pivot ──────────────────────────────────────────────────────────
    ("2026-04-25", "fd2c386", "Add AIR-7 cold-start fallback (v3) in case warm-start blocks"),
    ("2026-04-25", "84acd6a", "validator: add AIR-7 species branch (e-, N2, O2, NO, N, O, NO+)"),
    ("2026-04-25", "e290882", "AIR-7 v4: bump CFL to 0.5, target -2 convergence (hung in BCGSTAB)"),
    ("2026-04-25", "1ac4445", "AIR-7 v5: CFL=0.3 + LINEAR_SOLVER_ITER=15 (hung in BCGSTAB)"),
    ("2026-04-25", "1f7447c", "AIR-7 v6: revert to v3 settings (CFL=0.2) — CFL_ADAPT misfired"),
    ("2026-04-25", "ed89d0d", "AIR-7 v7: disable CFL_ADAPT (was killing iter rate)"),
    ("2026-04-25", "fcb9bdd", "AIR-7: skip M10 stage 1 grind, resume from M10 iter-200 checkpoint"),
    # ── AIR-7 result-prep tooling (while v7 cooks) ──────────────────────────
    ("2026-04-25", "ce1dd79", "Add AIR-7 result-prep tooling: comparison + chemistry plot + 3D diagnostic"),
    # ── MPI rebuild delegated to other instance ────────────────────────────
    ("2026-04-25", "(parallel)", "F-9: SU2-NEMO MPI rebuild kicked off on other Claude instance, separate /opt/su2-nemo-mpi/ prefix, expected 10-15x speedup"),
]


def write_tasks_sheet(ws):
    headers = [
        "Code", "Task", "Phase", "Status", "% Complete",
        "Original Timeline", "Started", "Finished / ETA",
        "Depends On", "Notes / Recent Update",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER

    for row_idx, t in enumerate(TASKS, 2):
        for col_idx, val in enumerate(t, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = BORDER
            c.alignment = LEFT if col_idx in (2, 10) else CENTER
        # Status fill
        ws.cell(row=row_idx, column=4).fill = fill_for_status(t[3])
        # % complete fill: green tint scaled
        pct = t[4]
        pct_cell = ws.cell(row=row_idx, column=5)
        pct_cell.value = f"{pct}%"
        if pct == 100: pct_cell.fill = GREEN
        elif pct >= 50: pct_cell.fill = YELLOW
        elif pct > 0:   pct_cell.fill = BLUE
        else:           pct_cell.fill = GRAY

    widths = [8, 50, 14, 14, 11, 18, 13, 16, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(TASKS) + 2):
        ws.row_dimensions[r].height = 38
    ws.freeze_panes = "B2"


def write_milestones_sheet(ws):
    headers = ["Week", "Milestone", "Original Target", "Actual / Projected",
               "Status", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER

    for row_idx, m in enumerate(MILESTONES, 2):
        for col_idx, val in enumerate(m, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = BORDER
            c.alignment = LEFT if col_idx == 6 else CENTER
        ws.cell(row=row_idx, column=5).fill = fill_for_status(m[4])

    widths = [13, 50, 16, 18, 14, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(MILESTONES) + 2):
        ws.row_dimensions[r].height = 36
    ws.freeze_panes = "B2"


def write_updates_sheet(ws):
    headers = ["Date", "Commit / Event", "Description"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER

    for row_idx, u in enumerate(UPDATES, 2):
        for col_idx, val in enumerate(u, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = BORDER
            c.alignment = LEFT if col_idx == 3 else CENTER

    widths = [14, 16, 100]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_summary_sheet(ws):
    """Top-level summary at-a-glance."""
    ws["A1"] = "Khorium PlasmaNet — Roadmap Status"
    ws["A1"].font = Font(bold=True, size=16, color="305496")
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Generated: {TODAY.isoformat()} | Day {(TODAY - START).days + 1} of project"
    ws["A2"].font = Font(italic=True, size=10, color="595959")
    ws.merge_cells("A2:D2")

    # Counts
    done = sum(1 for t in TASKS if t[3].lower().startswith("done"))
    in_prog = sum(1 for t in TASKS if "progress" in t[3].lower())
    pending = sum(1 for t in TASKS if t[3].lower().startswith("pending"))
    future = sum(1 for t in TASKS if t[3].lower().startswith("future"))
    total = len(TASKS)

    summary_rows = [
        ("",),
        ("Status", "Count", "% of Total"),
        ("Done",        done,    f"{done/total*100:.0f}%"),
        ("In Progress", in_prog, f"{in_prog/total*100:.0f}%"),
        ("Pending",     pending, f"{pending/total*100:.0f}%"),
        ("Future",      future,  f"{future/total*100:.0f}%"),
        ("TOTAL",       total,   "100%"),
    ]
    for r_off, row in enumerate(summary_rows, 4):
        for c_off, val in enumerate(row, 1):
            c = ws.cell(row=r_off, column=c_off, value=val)
            if r_off == 5:
                c.fill = HEADER; c.font = HEADER_FONT
                c.alignment = CENTER
            elif val in ("Done", "In Progress", "Pending", "Future"):
                c.fill = fill_for_status(val)
                c.alignment = LEFT
            else:
                c.alignment = CENTER

    ws["A12"] = "Critical path right now"
    ws["A12"].font = Font(bold=True, size=12)
    ws.merge_cells("A12:D12")

    crit = [
        "1. C-3a (refined-mesh ramp) — M18 stage running, M22.5 in queue. ETA ~5h to result.",
        "2. C-3b (validate refined M22.5) — fires automatically when C-3a finishes.",
        "3. C-3c (AIR-11 fallback) — pre-staged; only fires if log10 err > 1 after C-3b.",
        "4. C-4 (40-case batch port) — unblocked by C-3 success.",
        "5. I-5a (frontend selectors + 2nd chart) — in other-instance queue; demo polish.",
    ]
    for i, line in enumerate(crit, 13):
        ws.cell(row=i, column=1, value=line).alignment = LEFT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    ws["A20"] = "Key wins this week"
    ws["A20"].font = Font(bold=True, size=12)
    wins = [
        "✓ SU2-NEMO unblocked Day 1 (was 'doesn't work for 2 weeks' going in)",
        "✓ First RAM-C M22.5 NEMO result: log10 err +1.08 (vs +1.92 with equilibrium)",
        "✓ All 3 published BLACKOUT bands (VHF 225/450, X-band) match J&C qualitatively",
        "✓ Full SimOps CDK scaffold (Service + Worker stacks, cdk synth clean)",
        "✓ 71 contract tests (mock + lambda + PDF + agent) all green in CI",
        "✓ Frontend Storybook + vitest RTL coverage shipped (5 separate commits)",
        "✓ Repo cleanup: 1068 lines of dead code deleted (zero callers verified)",
        "✓ Refined mesh ramp running, ETA M22.5 result ~5h",
    ]
    for i, line in enumerate(wins, 21):
        ws.cell(row=i, column=1, value=line).alignment = LEFT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    widths = [50, 12, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_linear_import_sheet(ws):
    """Linear-friendly column layout — copy this whole sheet into Linear's import.

    Linear CSV/XLSX import expects columns:
      Title (required), Description, Status, Priority, Estimate, Labels.
    Map our internal codes to Linear's enum values:
      Status      -> Backlog | Todo | In Progress | Done | Canceled
      Priority    -> Urgent | High | Medium | Low | No priority
      Estimate    -> integer (rough hour count of focused dev time)
      Labels      -> comma-separated (Phase + any ad-hoc tags)
    """
    headers = [
        "Title", "Description", "Status", "Priority",
        "Estimate (hours)", "Labels", "Depends On", "Internal Code",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER

    status_map = {
        "done": "Done",
        "in progress": "In Progress",
        "pending": "Todo",
        "future": "Backlog",
        "blocked": "Backlog",
    }

    def priority_for(code: str, status: str) -> str:
        if status.lower().startswith("done"):
            return "No priority"
        # Validation-fix gap-closers
        if code in ("F-1", "F-7"):
            return "High"
        if code == "F-2":
            return "High"
        if code.startswith("F-"):
            return "Medium"
        if code.startswith("C-") or code.startswith("I-"):
            return "Medium"
        if code.startswith("P-"):
            return "Low"
        return "No priority"

    def effort_for(code: str) -> int:
        # Rough focused-work hours (NOT wall-clock; CFD runtime is in description)
        mapping = {
            "F-1": 4, "F-2": 6, "F-3": 8, "F-4": 2,
            "F-5": 6, "F-6": 2, "F-7": 2,
        }
        if code in mapping:
            return mapping[code]
        if code.startswith("H-"): return 0   # historical
        if code.startswith("C-"): return 8
        if code.startswith("I-"): return 8
        if code.startswith("T-"): return 16
        if code.startswith("P-"): return 12
        return 4

    for row_idx, t in enumerate(TASKS, 2):
        code, title, phase, status, pct, _orig, _start, _end, deps, notes = t
        linear_status = status_map.get(status.lower(), "Backlog")
        linear_priority = priority_for(code, status)
        linear_estimate = effort_for(code)
        labels = phase

        description_lines = [notes]
        if deps and deps != "—":
            description_lines.append(f"\nDepends on: {deps}")
        description_lines.append(f"\nInternal code: {code} | Phase: {phase}")
        if 0 < pct < 100:
            description_lines.append(f"Progress: {pct}%")
        description = "\n".join(description_lines)

        row_vals = [
            title,
            description,
            linear_status,
            linear_priority,
            linear_estimate,
            labels,
            deps,
            code,
        ]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = BORDER
            c.alignment = LEFT if col_idx in (1, 2, 6) else CENTER
        ws.cell(row=row_idx, column=3).fill = fill_for_status(status)

    widths = [40, 80, 14, 14, 12, 22, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(TASKS) + 2):
        ws.row_dimensions[r].height = 60
    ws.freeze_panes = "B2"


def main():
    wb = Workbook()

    # Reuse the default sheet for Summary
    summary = wb.active
    summary.title = "Summary"
    write_summary_sheet(summary)

    tasks = wb.create_sheet("Tasks")
    write_tasks_sheet(tasks)

    ms = wb.create_sheet("Milestones")
    write_milestones_sheet(ms)

    upd = wb.create_sheet("Recent Updates")
    write_updates_sheet(upd)

    linear = wb.create_sheet("Linear Import")
    write_linear_import_sheet(linear)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    out_path = _resolve_writable_path(OUT)
    wb.save(out_path)
    if out_path != OUT:
        print(f"Original {OUT.name} was locked (open in Excel?) — wrote {out_path.name} instead.")
    print(f"Wrote {out_path}")
    print(f"  Summary: counts + critical path + wins")
    print(f"  Tasks:   {len(TASKS)} rows, color-coded by status")
    print(f"  Milestones: {len(MILESTONES)} rows")
    print(f"  Recent Updates: {len(UPDATES)} commits/events")


if __name__ == "__main__":
    main()
